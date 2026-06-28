import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load data
df = pd.read_csv("data/customers_segmented.csv")

# Segment order and colours
seg_order = ["Premium Customers", "Impulsive Spenders",
             "Standard Customers", "Cautious Savers", "Budget Customers"]

seg_colors = {
    "Premium Customers":   "1B3A6B",
    "Impulsive Spenders":  "C9A84C",
    "Standard Customers":  "0D6E6E",
    "Cautious Savers":     "C0392B",
    "Budget Customers":    "6C3483"
}

# Helper functions
def hfill(hex): return PatternFill("solid", fgColor=hex)
def hfont(bold=True, color="FFFFFF", size=11):
    return Font(name="Arial", bold=bold, color=color, size=size)
def bfont(bold=False, color="1A1A2E", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)
def center(): return Alignment(horizontal="center", vertical="center")
def thin_border():
    s = Side(border_style="thin", color="D0D5E0")
    return Border(left=s, right=s, top=s, bottom=s)

wb = Workbook()
wb.remove(wb.active)

print("Setup done — ready to build sheets.")

# ── SHEET 1: COVER ──────────────────────────────────────────────────────────
ws1 = wb.create_sheet("Cover")
ws1.sheet_view.showGridLines = False

# Column widths
for col in "ABCDE":
    ws1.column_dimensions[col].width = 22

# Title block
ws1.merge_cells("A1:E6")
ws1["A1"] = "CUSTOMER SEGMENTATION ANALYSIS"
ws1["A1"].font = Font(name="Arial", bold=True, size=28, color="FFFFFF")
ws1["A1"].fill = hfill("1B3A6B")
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 80

for r in range(2, 7):
    ws1.merge_cells(f"A{r}:E{r}")
    ws1[f"A{r}"].fill = hfill("1B3A6B")

# Subtitle
ws1.merge_cells("A7:E7")
ws1["A7"] = "K-Means Clustering  |  Mall Customer Dataset  |  Python & Scikit-learn"
ws1["A7"].font = Font(name="Arial", size=12, color="C9A84C", italic=True)
ws1["A7"].fill = hfill("1B3A6B")
ws1["A7"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[7].height = 30

# Gold divider
ws1.merge_cells("A8:E8")
ws1["A8"].fill = hfill("C9A84C")
ws1.row_dimensions[8].height = 4

# KPI cards
kpis = [
    ("200",   "Customers"),
    ("5",     "Segments"),
    ("0.355", "Silhouette"),
    ("40.5%", "Largest Seg"),
    ("19.5%", "Premium Seg"),
]
cards_row1 = 10
cards_row2 = 11
col_letters = ["A", "B", "C", "D", "E"]
card_colors = ["1B3A6B", "0D6E6E", "C9A84C", "C0392B", "6C3483"]

ws1.row_dimensions[cards_row1].height = 45
ws1.row_dimensions[cards_row2].height = 22

for i, (val, label) in enumerate(kpis):
    c = col_letters[i]
    ws1[f"{c}{cards_row1}"] = val
    ws1[f"{c}{cards_row1}"].font = Font(name="Arial", bold=True, size=22, color="FFFFFF")
    ws1[f"{c}{cards_row1}"].fill = hfill(card_colors[i])
    ws1[f"{c}{cards_row1}"].alignment = Alignment(horizontal="center", vertical="center")

    ws1[f"{c}{cards_row2}"] = label
    ws1[f"{c}{cards_row2}"].font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws1[f"{c}{cards_row2}"].fill = hfill(card_colors[i])
    ws1[f"{c}{cards_row2}"].alignment = Alignment(horizontal="center", vertical="center")

# Sheet index
ws1.row_dimensions[13].height = 26
ws1.merge_cells("A13:E13")
ws1["A13"] = "WORKBOOK CONTENTS"
ws1["A13"].font = Font(name="Arial", bold=True, size=12, color="1B3A6B")

sheets_list = [
    ("Cover",           "Project overview and key metrics"),
    ("Raw Data",        "Original Mall Customers dataset"),
    ("Segment Profiles","Mean feature values per segment"),
    ("Segmented Data",  "All 200 customers with segment labels"),
    ("Recommendations", "Marketing strategy per segment"),
]
for i, (sh, desc) in enumerate(sheets_list):
    r = 14 + i
    ws1.row_dimensions[r].height = 22
    bg = "1B3A6B" if i % 2 == 0 else "2E4F8A"
    ws1[f"A{r}"] = sh
    ws1[f"A{r}"].font = hfont(size=10)
    ws1[f"A{r}"].fill = hfill(bg)
    ws1[f"A{r}"].alignment = center()
    ws1.merge_cells(f"B{r}:E{r}")
    ws1[f"B{r}"] = desc
    ws1[f"B{r}"].font = hfont(bold=False, size=10)
    ws1[f"B{r}"].fill = hfill(bg)
    ws1[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")

# Footer
ws1.row_dimensions[20].height = 22
ws1.merge_cells("A20:E20")
ws1["A20"] = "Arun  |  Data Analyst  |  github.com/arunofficial25"
ws1["A20"].font = Font(name="Arial", size=9, color="999999", italic=True)
ws1["A20"].alignment = Alignment(horizontal="center", vertical="center")

print("Sheet 1 (Cover) done.")

# ── SHEET 2: RAW DATA ───────────────────────────────────────────────────────
ws2 = wb.create_sheet("Raw Data")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A2"

headers = ["Customer ID", "Gender", "Age", "Annual Income (k$)", "Spending Score (1-100)"]
col_widths = [13, 10, 8, 18, 22]

# Header row
for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws2.cell(1, ci, h)
    cell.font      = hfont(size=10)
    cell.fill      = hfill("1B3A6B")
    cell.alignment = center()
    cell.border    = thin_border()
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[1].height = 28

# Data rows
raw_cols = ["CustomerID", "Gender", "Age", "Annual Income (k$)", "Spending Score (1-100)"]
for ri, row in enumerate(df[raw_cols].itertuples(index=False), 2):
    bg = "FFFFFF" if ri % 2 == 0 else "F2F4F7"
    for ci, val in enumerate(row, 1):
        cell = ws2.cell(ri, ci, val)
        cell.font      = bfont(size=9)
        cell.fill      = hfill(bg)
        cell.alignment = center()
        cell.border    = thin_border()
    ws2.row_dimensions[ri].height = 18

print("Sheet 2 (Raw Data) done.")

# ── SHEET 3: SEGMENT PROFILES ───────────────────────────────────────────────
ws3 = wb.create_sheet("Segment Profiles")
ws3.sheet_view.showGridLines = False

# Title
ws3.merge_cells("A1:F1")
ws3["A1"] = "SEGMENT PROFILES — K-MEANS (K=5)"
ws3["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws3["A1"].fill = hfill("1B3A6B")
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 30

# Headers
prof_headers = ["Segment", "Customers", "Share (%)", "Avg Age",
                "Avg Income (k$)", "Avg Spending Score"]
prof_widths  = [22, 12, 11, 10, 17, 20]

for ci, (h, w) in enumerate(zip(prof_headers, prof_widths), 1):
    cell = ws3.cell(2, ci, h)
    cell.font      = hfont(size=10)
    cell.fill      = hfill("0D6E6E")
    cell.alignment = center()
    cell.border    = thin_border()
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.row_dimensions[2].height = 26

# Profile data
profile = df.groupby("Segment").agg(
    Count=("Segment", "count"),
    Avg_Age=("Age", "mean"),
    Avg_Income=("Annual Income (k$)", "mean"),
    Avg_Spending=("Spending Score (1-100)", "mean")
).round(1).reset_index()
profile["Share"] = (profile["Count"] / len(df) * 100).round(1)
profile["Segment"] = pd.Categorical(profile["Segment"], seg_order)
profile = profile.sort_values("Segment").reset_index(drop=True)

for ri, row in profile.iterrows():
    rn = ri + 3
    seg = row["Segment"]
    color = seg_colors[seg]
    vals = [seg, int(row["Count"]), f'{row["Share"]}%',
            row["Avg_Age"], row["Avg_Income"], row["Avg_Spending"]]
    for ci, val in enumerate(vals, 1):
        cell = ws3.cell(rn, ci, val)
        if ci == 1:
            cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill = hfill(color)
        else:
            cell.font = bfont(size=10)
            cell.fill = hfill("F8F9FA" if ri % 2 == 0 else "FFFFFF")
        cell.alignment = center()
        cell.border    = thin_border()
    ws3.row_dimensions[rn].height = 24

# Insight box
ws3.row_dimensions[10].height = 10
ws3.merge_cells("A11:F11")
ws3["A11"] = "KEY INSIGHT"
ws3["A11"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws3["A11"].fill = hfill("C9A84C")
ws3["A11"].alignment = center()
ws3.row_dimensions[11].height = 24

insights = [
    "Premium Customers (19.5%) — High income, high spend. Top priority for retention & upsell.",
    "Impulsive Spenders (11%) — Young, low income but high spend. Respond well to flash sales.",
    "Standard Customers (40.5%) — Largest group. Nurture with loyalty programmes.",
    "Cautious Savers (17.5%) — High income, low spend. Untapped potential — needs trust building.",
    "Budget Customers (11.5%) — Price-sensitive. Target with discounts and value offers.",
]
for i, text in enumerate(insights):
    r = 12 + i
    ws3.merge_cells(f"A{r}:F{r}")
    ws3[f"A{r}"] = text
    ws3[f"A{r}"].font = bfont(size=9)
    ws3[f"A{r}"].fill = hfill("FFFDE7" if i % 2 == 0 else "FFFFFF")
    ws3[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
    ws3[f"A{r}"].border = thin_border()
    ws3.row_dimensions[r].height = 20

print("Sheet 3 (Segment Profiles) done.")

# ── SHEET 4: SEGMENTED DATA ─────────────────────────────────────────────────
ws4 = wb.create_sheet("Segmented Data")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A2"

seg_headers = ["Customer ID", "Gender", "Age",
               "Annual Income (k$)", "Spending Score (1-100)", "Segment"]
seg_widths   = [13, 10, 8, 18, 22, 22]

for ci, (h, w) in enumerate(zip(seg_headers, seg_widths), 1):
    cell = ws4.cell(1, ci, h)
    cell.font      = hfont(size=10)
    cell.fill      = hfill("1B3A6B")
    cell.alignment = center()
    cell.border    = thin_border()
    ws4.column_dimensions[get_column_letter(ci)].width = w
ws4.row_dimensions[1].height = 28

seg_cols = ["CustomerID", "Gender", "Age",
            "Annual Income (k$)", "Spending Score (1-100)", "Segment"]
for ri, row in enumerate(df[seg_cols].itertuples(index=False), 2):
    seg   = row[5]
    color = seg_colors[seg]
    bg    = "FFFFFF" if ri % 2 == 0 else "F2F4F7"
    for ci, val in enumerate(row, 1):
        cell = ws4.cell(ri, ci, val)
        if ci == 6:
            cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            cell.fill = hfill(color)
        else:
            cell.font = bfont(size=9)
            cell.fill = hfill(bg)
        cell.alignment = center()
        cell.border    = thin_border()
    ws4.row_dimensions[ri].height = 18

print("Sheet 4 (Segmented Data) done.")

# ── SHEET 5: RECOMMENDATIONS ────────────────────────────────────────────────
ws5 = wb.create_sheet("Recommendations")
ws5.sheet_view.showGridLines = False

# Title
ws5.merge_cells("A1:E1")
ws5["A1"] = "MARKETING RECOMMENDATIONS BY SEGMENT"
ws5["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
ws5["A1"].fill = hfill("1B3A6B")
ws5["A1"].alignment = center()
ws5.row_dimensions[1].height = 30

# Headers
rec_headers = ["Segment", "Strategy", "Tactics", "Goal", "Channel"]
rec_widths  = [22, 20, 40, 25, 20]

for ci, (h, w) in enumerate(zip(rec_headers, rec_widths), 1):
    cell = ws5.cell(2, ci, h)
    cell.font      = hfont(size=10)
    cell.fill      = hfill("0D6E6E")
    cell.alignment = center()
    cell.border    = thin_border()
    ws5.column_dimensions[get_column_letter(ci)].width = w
ws5.row_dimensions[2].height = 26

# Recommendations data
recs = [
    ("Premium Customers",  "Retain & Upsell",     "Loyalty rewards, exclusive offers, early access to new products", "Maximise CLV",          "Email, App"),
    ("Impulsive Spenders", "Capitalise & Convert", "Flash sales, limited-time deals, social media campaigns",        "Increase basket size",  "Instagram, SMS"),
    ("Standard Customers", "Nurture & Grow",       "Loyalty programme, personalised recommendations, newsletters",   "Move to Premium tier",  "Email, In-store"),
    ("Cautious Savers",    "Build Trust",          "Highlight value, quality content, money-back guarantees",        "Unlock spending",       "Blog, Email"),
    ("Budget Customers",   "Value Offers",         "Discount coupons, bundle deals, seasonal promotions",           "Drive repeat purchase", "SMS, Flyers"),
]

for ri, (seg, strategy, tactics, goal, channel) in enumerate(recs):
    rn    = ri + 3
    color = seg_colors[seg]
    bg    = "F8F9FA" if ri % 2 == 0 else "FFFFFF"
    vals  = [seg, strategy, tactics, goal, channel]

    for ci, val in enumerate(vals, 1):
        cell = ws5.cell(rn, ci, val)
        if ci == 1:
            cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill = hfill(color)
        else:
            cell.font = bfont(size=9)
            cell.fill = hfill(bg)
        cell.alignment = Alignment(horizontal="left" if ci >= 3 else "center",
                                   vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws5.row_dimensions[rn].height = 45

print("Sheet 5 (Recommendations) done.")

# ── SAVE WORKBOOK ────────────────────────────────────────────────────────────
wb.save("outputs/Customer_Segmentation_Report.xlsx")
print("\nWorkbook saved → outputs/Customer_Segmentation_Report.xlsx")
